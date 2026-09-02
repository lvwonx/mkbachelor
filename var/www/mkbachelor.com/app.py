from flask import Flask, request, jsonify, render_template, session, redirect
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg
from functools import wraps
from psycopg.rows import dict_row
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
import io
from io import BytesIO
from datetime import datetime
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from psycopg.types.json import Jsonb
import json
import os
from dotenv import load_dotenv
import traceback

load_dotenv("/var/www/mkbachelor.com/.env")

app = Flask(__name__)
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = True
app.secret_key = os.environ["FLASK_SECRET_KEY"]
pdfmetrics.registerFont(
    TTFont(
        "DejaVuSans",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    )
)

def get_db_connection():
    return psycopg.connect(
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        host=os.environ["DB_HOST"],
    )
    
def write_audit_log(
    cursor,
    action,
    entity_type,
    entity_id=None,
    details=None
):
    user_id = session.get("user_id")

    cursor.execute("""
        INSERT INTO audit_log (
            user_id,
            action,
            entity_type,
            entity_id,
            details
        )
        VALUES (%s, %s, %s, %s, %s)
    """, (
        user_id,
        action,
        entity_type,
        entity_id,
        Jsonb(details) if details is not None else None
    ))

    try:
        username = session.get("username", "-")

        timestamp = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        details_text = (
            json.dumps(
                details,
                ensure_ascii=False
            )
            if details is not None
            else "-"
        )

        log_line = (
            f"[{timestamp}] "
            f"USER={username} | "
            f"ACTION={action} | "
            f"ENTITY={entity_type} | "
            f"ENTITY_ID={entity_id if entity_id is not None else '-'} | "
            f"DETAILS={details_text}\n"
        )

        with open(
            "/var/www/mkbachelor.com/audit.log",
            "a",
            encoding="utf-8"
        ) as log_file:
            log_file.write(log_line)

    except Exception as e:
        print("AUDIT FILE ERROR:", str(e))

def is_logged_in():
    return "user_id" in session
    
def has_permission(permission):
    if "user_id" not in session:
        return False

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT u.session_version
        FROM users u
        WHERE u.user_id = %s
    """, (session["user_id"],))

    user = cur.fetchone()

    if not user:
        cur.close()
        conn.close()
        return False

    if session.get("session_version") != user[0]:
        cur.close()
        conn.close()
        session.clear()
        return False

    cur.execute("""
        SELECT 1
        FROM users u
        JOIN roles r ON u.role_id = r.id
        JOIN role_permissions rp ON rp.role_id = r.id
        JOIN permissions p ON rp.permission_id = p.id
        WHERE u.user_id = %s AND p.name = %s
    """, (session["user_id"], permission))

    result = cur.fetchone()

    cur.close()
    conn.close()

    return result is not None

def require_permission(permission):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not has_permission(permission):
                return {"error": "Forbidden"}, 403
            return f(*args, **kwargs)

        return wrapper

    return decorator

@app.route("/")
def dashboard():
    if not is_logged_in():
        return redirect("/login")

    return render_template("dashboard.html", username=session["username"])

@app.route("/register", methods=["GET"])
def register_page():
    return render_template("register.html")
    
@app.route("/login", methods=["GET"])
def login_page():
    return render_template("login.html")

@app.route("/risks-page")
@require_permission("view_event")
def risks_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template("risks.html", username=session["username"])

@app.route("/register", methods=["POST"])
def register():
    data = request.json

    username = data.get("username")
    password = data.get("password")

    if not username or not password:
        return {"error": "Missing fields"}, 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT 1 FROM users WHERE username = %s",(username,))

        if cursor.fetchone():
            return {"error": "User already exists"}, 400

        password_hash = generate_password_hash(password)

        cursor.execute("""
            INSERT INTO users (
                username,
                password_hash,
                role_id
            )
            VALUES (
                %s,
                %s,
                (SELECT id FROM roles WHERE name = 'guest')
            )
            RETURNING user_id
        """, (
            username,
            password_hash
        ))

        user_id = cursor.fetchone()[0]

        write_audit_log(
            cursor,
            "CREATE",
            "user",
            user_id,
            {
                "username": username,
                "role": "guest",
                "source": "registration"
            }
        )

        conn.commit()

        return jsonify({
            "message": "registered"
        })

    except Exception as e:
        conn.rollback()

        print("REGISTER ERROR:", str(e))

        return jsonify({
            "error": "Registration failed"
        }), 500

    finally:
        cursor.close()
        conn.close()

def check_login_rate_limit(ip_address):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO login_rate_limits (
                ip_address,
                window_start,
                attempts
            )
            VALUES (
                %s,
                NOW(),
                0
            )
            ON CONFLICT (ip_address)
            DO UPDATE SET
                window_start = CASE
                    WHEN NOW() - login_rate_limits.window_start >= INTERVAL '1 minute'
                    THEN NOW()
                    ELSE login_rate_limits.window_start
                END,
                attempts = CASE
                    WHEN NOW() - login_rate_limits.window_start >= INTERVAL '1 minute'
                    THEN 0
                    ELSE login_rate_limits.attempts
                END
            RETURNING attempts
        """, (ip_address,))

        attempts = cur.fetchone()[0]

        conn.commit()

        return attempts

    finally:
        cur.close()
        conn.close()
        
def record_login_failure(ip_address):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE login_rate_limits
            SET attempts = attempts + 1
            WHERE ip_address = %s
        """, (ip_address,))

        conn.commit()

    finally:
        cur.close()
        conn.close()

@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}

    username = data.get("username")
    password = data.get("password")

    client_ip = request.remote_addr

    attempts = check_login_rate_limit(client_ip)

    if attempts >= 10:
        return jsonify({
            "error": "Too many login attempts"
        }), 429, {
            "Retry-After": "60"
        }

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM users WHERE username=%s",(username,))

        user = cursor.fetchone()

        if not user or not check_password_hash(user[2], password):
            record_login_failure(client_ip)
            write_audit_log(
                cursor,
                "LOGIN_FAILED",
                "authentication",
                None,
                {
                    "username": username
                }
            )

            conn.commit()

            return jsonify({
                "error": "invalid"
            }), 401

        session["user_id"] = user[0]
        session["username"] = user[1]
        session["role_id"] = user[4]
        session["session_version"] = user[5]

        write_audit_log(
            cursor,
            "LOGIN",
            "authentication",
            user[0],
            {
                "username": user[1]
            }
        )

        conn.commit()

        return jsonify({
            "message": "ok"
        })

    except Exception as e:
        conn.rollback()

        print("LOGIN AUDIT ERROR:", str(e))

        return jsonify({
            "error": "Login failed"
        }), 500

    finally:
        cursor.close()
        conn.close()

@app.route("/add-risk", methods=["POST"])
@require_permission("create_event")
def add_risk():
    data = request.get_json(silent=True)

    if not data:
        return jsonify({"error": "Invalid JSON data"}), 400

    name = str(data.get("name", "")).strip()
    description = str(data.get("description", "")).strip()
    status = data.get("status", "active")

    try:
        probability = int(data.get("probability"))
        impact = int(data.get("impact"))
        system_id = int(data.get("system_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid numeric values"}), 400

    if not name:
        return jsonify({"error": "Risk name is required"}), 400

    if not 1 <= probability <= 5:
        return jsonify({
            "error": "Probability must be between 1 and 5"
        }), 400

    if not 1 <= impact <= 5:
        return jsonify({
            "error": "Impact must be between 1 and 5"
        }), 400

    if status not in {"active", "mitigated", "closed"}:
        return jsonify({"error": "Invalid status"}), 400

    risk_level = probability * impact

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT 1
            FROM systems
            WHERE system_id = %s
        """, (system_id,))

        if cursor.fetchone() is None:
            return jsonify({
                "error": "Information system not found"
            }), 404

        cursor.execute("""
            INSERT INTO risks (
                name,
                description,
                probability,
                impact,
                risk_level,
                status,
                system_id,
                created_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING risk_id
        """, (
            name,
            description,
            probability,
            impact,
            risk_level,
            status,
            system_id,
            session["user_id"]
        ))

        risk_id = cursor.fetchone()[0]

        write_audit_log(
            cursor,
            "CREATE",
            "risk",
            risk_id,
            {
                "name": name,
                "system_id": system_id,
                "risk_level": risk_level
            }
        )

        conn.commit()

        return jsonify({
            "message": "Risk added",
            "risk_id": risk_id,
            "risk_level": risk_level
        }), 201

    except Exception as e:
        conn.rollback()

        print("ADD RISK ERROR:", str(e))
        traceback.print_exc()

        return jsonify({
            "error": "Failed to add risk"
        }), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/risks", methods=["GET"])
@require_permission("view_event")
def get_risks():

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        page = request.args.get(
            "page",
            default=1,
            type=int
        )

        per_page = request.args.get(
            "per_page",
            default=5,
            type=int
        )

        search = request.args.get(
            "search",
            default="",
            type=str
        ).strip()

        if page < 1:
            page = 1

        if per_page not in (5, 10, 20, 50):
            per_page = 5

        offset = (page - 1) * per_page

        search_pattern = f"%{search}%"

        cursor.execute("""
            SELECT COUNT(*)

            FROM risks r

            LEFT JOIN users u
                ON r.created_by = u.user_id

            LEFT JOIN systems s
                ON r.system_id = s.system_id

            WHERE
                %s = ''
                OR r.name ILIKE %s
                OR r.description ILIKE %s
                OR r.status ILIKE %s
                OR CAST(r.risk_level AS TEXT) ILIKE %s
                OR s.name ILIKE %s
                OR u.username ILIKE %s
        """, (
            search,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern
        ))

        total = cursor.fetchone()[0]

        pages = (
            (total + per_page - 1)
            // per_page
            if total > 0
            else 1
        )

        if page > pages:
            page = pages
            offset = (page - 1) * per_page

        cursor.execute("""
            SELECT
                r.risk_id,
                r.name,
                r.description,
                r.probability,
                r.impact,
                r.risk_level,
                r.status,
                r.system_id,
                r.created_at,
                u.username,
                s.name

            FROM risks r

            LEFT JOIN users u
                ON r.created_by = u.user_id

            LEFT JOIN systems s
                ON r.system_id = s.system_id

            WHERE
                %s = ''
                OR r.name ILIKE %s
                OR r.description ILIKE %s
                OR r.status ILIKE %s
                OR CAST(r.risk_level AS TEXT) ILIKE %s
                OR s.name ILIKE %s
                OR u.username ILIKE %s

            ORDER BY r.created_at DESC

            LIMIT %s
            OFFSET %s
        """, (
            search,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern,
            search_pattern,
            per_page,
            offset
        ))

        rows = cursor.fetchall()

        risks = []

        for row in rows:
            risks.append({
                "risk_id": row[0],
                "name": row[1],
                "description": row[2],
                "probability": row[3],
                "impact": row[4],
                "risk_level": row[5],
                "status": row[6],
                "system_id": row[7],
                "created_at":
                    str(row[8]),
                "created_by":
                    row[9],
                "system_name":
                    row[10]
            })

        return jsonify({
            "items": risks,
            "page": page,
            "pages": pages,
            "per_page": per_page,
            "total": total
        })

    except Exception as e:
        print(
            "GET RISKS ERROR:",
            str(e)
        )

        return jsonify({
            "error":
                "Failed to load risks"
        }), 500

    finally:
        cursor.close()
        conn.close()
        
@app.route("/risks/statistics", methods=["GET"])
@require_permission("view_event")
def get_risk_statistics():
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                COUNT(*) FILTER (
                    WHERE risk_level <= 5
                ) AS low_count,

                COUNT(*) FILTER (
                    WHERE risk_level > 5
                    AND risk_level <= 12
                ) AS medium_count,

                COUNT(*) FILTER (
                    WHERE risk_level > 12
                ) AS high_count

            FROM risks
        """)

        result = cursor.fetchone()

        return jsonify({
            "low": result[0],
            "medium": result[1],
            "high": result[2]
        })

    except Exception as error:
        print(
            "GET RISK STATISTICS ERROR:",
            error
        )

        return jsonify({
            "error":
                "Failed to load risk statistics"
        }), 500

    finally:
        cursor.close()
        conn.close()

@app.route("/risks/<int:risk_id>", methods=["GET"])
@require_permission("edit_event")
def get_risk(risk_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                risk_id,
                name,
                description,
                probability,
                impact,
                risk_level,
                status,
                system_id
            FROM risks
            WHERE risk_id = %s
        """, (risk_id,))

        row = cursor.fetchone()

        if not row:
            return jsonify({"error": "Risk not found"}), 404

        return jsonify({
            "risk_id": row[0],
            "name": row[1],
            "description": row[2],
            "probability": row[3],
            "impact": row[4],
            "risk_level": row[5],
            "status": row[6],
            "system_id": row[7]
        })

    finally:
        cursor.close()
        conn.close()

@app.route("/risks/<int:risk_id>", methods=["PUT"])
@require_permission("edit_event")
def update_risk(risk_id):
    data = request.get_json(silent=True)

    if not data:
        return jsonify({"error": "Invalid JSON data"}), 400

    name = str(data.get("name", "")).strip()
    description = str(data.get("description", "")).strip()
    status = data.get("status")

    try:
        probability = int(data.get("probability"))
        impact = int(data.get("impact"))
        system_id = int(data.get("system_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid numeric values"}), 400

    if not name:
        return jsonify({"error": "Risk name is required"}), 400

    if not 1 <= probability <= 5:
        return jsonify({
            "error": "Probability must be between 1 and 5"
        }), 400

    if not 1 <= impact <= 5:
        return jsonify({
            "error": "Impact must be between 1 and 5"
        }), 400

    if status not in {"active", "mitigated", "closed"}:
        return jsonify({"error": "Invalid status"}), 400

    risk_level = probability * impact

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT 1
            FROM systems
            WHERE system_id = %s
        """, (system_id,))

        if cursor.fetchone() is None:
            return jsonify({
                "error": "Information system not found"
            }), 404

        cursor.execute("""
            UPDATE risks
            SET
                name = %s,
                description = %s,
                probability = %s,
                impact = %s,
                risk_level = %s,
                status = %s,
                system_id = %s
            WHERE risk_id = %s
        """, (
            name,
            description,
            probability,
            impact,
            risk_level,
            status,
            system_id,
            risk_id
        ))

        if cursor.rowcount == 0:
            conn.rollback()
            return jsonify({"error": "Risk not found"}), 404

        write_audit_log(
            cursor,
            "UPDATE",
            "risk",
            risk_id,
            {
                "name": name,
                "system_id": system_id,
                "risk_level": risk_level,
                "status": status
            }
        )

        conn.commit()

        return jsonify({
            "message": "Risk updated",
            "risk_level": risk_level
        })

    except Exception as e:
        conn.rollback()
        print("UPDATE RISK ERROR:", str(e))

        return jsonify({
            "error": "Failed to update risk"
        }), 500

    finally:
        cursor.close()
        conn.close()

@app.route("/risks/<int:risk_id>", methods=["DELETE"])
@require_permission("delete_event")
def delete_risk(risk_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT 1
            FROM risks
            WHERE risk_id = %s
        """, (risk_id,))

        if cursor.fetchone() is None:
            return jsonify({"error": "Risk not found"}), 404

        cursor.execute("""
            SELECT 1
            FROM incidents
            WHERE risk_id = %s
            LIMIT 1
        """, (risk_id,))

        if cursor.fetchone():
            return jsonify({
                "error": "Risk is used by an incident and cannot be deleted"
            }), 409

        cursor.execute("""
            DELETE FROM risks
            WHERE risk_id = %s
        """, (risk_id,))

        write_audit_log(
            cursor,
            "DELETE",
            "risk",
            risk_id
        )

        conn.commit()

        return jsonify({
            "message": "Risk deleted"
        })

        conn.commit()

        return jsonify({
            "message": "Risk deleted"
        })

    except Exception as e:
        conn.rollback()
        print("DELETE RISK ERROR:", str(e))

        return jsonify({
            "error": "Failed to delete risk"
        }), 500

    finally:
        cursor.close()
        conn.close()

@app.route("/systems", methods=["GET"])
@require_permission("view_system")
def get_systems():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = request.args.get(
            "per_page",
            5,
            type=int
        )

        search = request.args.get(
            "search",
            "",
            type=str
        ).strip()


        if page < 1:
            page = 1


        if per_page not in (
            5,
            10,
            20,
            50
        ):
            per_page = 5

        search_condition = ""
        search_params = []

        if search:
            search_condition = """
                WHERE
                    CAST(s.system_id AS TEXT)
                        ILIKE %s

                    OR s.name ILIKE %s

                    OR s.description ILIKE %s

                    OR u.username ILIKE %s

                    OR CAST(
                        s.cia_confidentiality
                        AS TEXT
                    ) ILIKE %s

                    OR CAST(
                        s.cia_integrity
                        AS TEXT
                    ) ILIKE %s

                    OR CAST(
                        s.cia_availability
                        AS TEXT
                    ) ILIKE %s

                    OR CAST(
                        s.classification
                        AS TEXT
                    ) ILIKE %s

                    OR s.security_level ILIKE %s
            """

            search_value = f"%{search}%"

            search_params = [
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value
            ]

        cur.execute(
            f"""
            SELECT COUNT(*)

            FROM systems s

            LEFT JOIN users u
                ON s.owner_id = u.user_id

            {search_condition}
            """,
            search_params
        )

        total = cur.fetchone()[0]

        pages = (
            (total + per_page - 1)
            // per_page
        )

        if pages == 0:

            page = 1

        elif page > pages:

            page = pages

        offset = (
            (page - 1)
            * per_page
        )

        cur.execute(
            f"""
            SELECT
                s.system_id,
                s.name,
                s.description,
                u.username,
                s.owner_id,
                s.cia_confidentiality,
                s.cia_integrity,
                s.cia_availability,
                s.classification,
                s.security_level

            FROM systems s

            LEFT JOIN users u
                ON s.owner_id = u.user_id

            {search_condition}

            ORDER BY
                s.system_id DESC

            LIMIT %s
            OFFSET %s
            """,
            search_params + [
                per_page,
                offset
            ]
        )

        rows = cur.fetchall()

        result = []

        for s in rows:
            result.append({
                "id": s[0],
                "name": s[1],
                "description": s[2],
                "owner": s[3],
                "owner_id": s[4],
                "cia_confidentiality": s[5],
                "cia_integrity": s[6],
                "cia_availability": s[7],
                "classification": s[8],
                "security_level": s[9]
            })

        return jsonify({
            "items":
                result,
            "page":
                page,
            "pages":
                pages,
            "per_page":
                per_page,
            "total":
                total
        })

    except Exception as e:
        print(
            "GET SYSTEMS ERROR:",
            str(e)
        )

        return jsonify({
            "error":
                "Failed to load systems"
        }), 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/systems-page")
@require_permission("view_system")
def systems_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template("systems.html", username=session["username"])

@app.route("/api/users/list", methods=["GET"])
@require_permission("view_system")
def get_users_list():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT user_id, username FROM users")

    users = [
        {"id": u[0], "username": u[1]}
        for u in cur.fetchall()
    ]

    cur.close()
    conn.close()

    return jsonify(users)

@app.route("/systems/<int:system_id>/history", methods=["GET"])
@require_permission("view_system")
def get_system_classification_history(system_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            h.id,
            h.cia_confidentiality,
            h.cia_integrity,
            h.cia_availability,
            h.classification,
            h.security_level,
            u.username,
            h.changed_at
        FROM system_classification_history h
        LEFT JOIN users u
            ON h.changed_by = u.user_id
        WHERE h.system_id = %s
        ORDER BY h.changed_at DESC
    """, (system_id,))

    rows = cur.fetchall()

    result = []

    for h in rows:
        result.append({
            "id": h[0],
            "cia_confidentiality": h[1],
            "cia_integrity": h[2],
            "cia_availability": h[3],
            "classification": h[4],
            "security_level": h[5],
            "changed_by": h[6],
            "changed_at": str(h[7]) if h[7] else None
        })

    cur.close()
    conn.close()

    return jsonify(result)

@app.route("/systems", methods=["POST"])
@require_permission("create_system")
def add_system():
    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cia_confidentiality = int(data.get("cia_confidentiality"))
        cia_integrity = int(data.get("cia_integrity"))
        cia_availability = int(data.get("cia_availability"))

        classification = max(
            cia_confidentiality,
            cia_integrity,
            cia_availability
        )

        if classification <= 2:
            security_level = "Zems"
        elif classification == 3:
            security_level = "Vidējs"
        else:
            security_level = "Augsts"

        cur.execute("""
            INSERT INTO systems (
                name,
                description,
                owner_id,
                cia_confidentiality,
                cia_integrity,
                cia_availability,
                classification,
                security_level
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING system_id
        """, (
            data.get("name"),
            data.get("description"),
            data.get("owner_id"),
            cia_confidentiality,
            cia_integrity,
            cia_availability,
            classification,
            security_level
        ))

        system_id = cur.fetchone()[0]

        cur.execute("""
            INSERT INTO system_classification_history (
                system_id,
                cia_confidentiality,
                cia_integrity,
                cia_availability,
                classification,
                security_level,
                changed_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            system_id,
            cia_confidentiality,
            cia_integrity,
            cia_availability,
            classification,
            security_level,
            session["user_id"]
        ))
        
        write_audit_log(
            cur,
            "CREATE",
            "system",
            system_id,
            {
                "name": data.get("name"),
                "owner_id": data.get("owner_id"),
                "classification": classification,
                "security_level": security_level
            }
        )

        conn.commit()

        return {
            "message": "System created",
            "classification": classification,
            "security_level": security_level
        }

    except Exception as e:
        conn.rollback()

        print("CREATE SYSTEM ERROR:", str(e))
        traceback.print_exc()

        return jsonify({
            "error": "Failed to create system",
        }), 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/systems/<int:system_id>", methods=["DELETE"])
@require_permission("delete_system")
def delete_system(system_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT 1 FROM systems WHERE system_id = %s", (system_id,))
    if not cur.fetchone():
        return {"error": "System not found"}, 404

    cur.execute("DELETE FROM systems WHERE system_id = %s", (system_id,))
    
    write_audit_log(
        cur,
        "DELETE",
        "system",
        system_id
    )

    conn.commit()

    cur.close()
    conn.close()

    return {"message": "System deleted"}
    
@app.route("/systems/<int:system_id>", methods=["PUT"])
@require_permission("edit_system")
def update_system(system_id):
    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            cia_confidentiality,
            cia_integrity,
            cia_availability
        FROM systems
        WHERE system_id = %s
    """, (system_id,))

    old_data = cur.fetchone()

    if not old_data:
        cur.close()
        conn.close()
        return {"error": "System not found"}, 404

    cia_confidentiality = int(data.get("cia_confidentiality"))
    cia_integrity = int(data.get("cia_integrity"))
    cia_availability = int(data.get("cia_availability"))

    classification = max(
        cia_confidentiality,
        cia_integrity,
        cia_availability
    )

    if classification <= 2:
        security_level = "Zems"
    elif classification == 3:
        security_level = "Vidējs"
    else:
        security_level = "Augsts"

    cia_changed = (
        old_data[0] != cia_confidentiality or
        old_data[1] != cia_integrity or
        old_data[2] != cia_availability
    )

    cur.execute("""
        UPDATE systems
        SET
            name = %s,
            description = %s,
            owner_id = %s,
            cia_confidentiality = %s,
            cia_integrity = %s,
            cia_availability = %s,
            classification = %s,
            security_level = %s
        WHERE system_id = %s
    """, (
        data.get("name"),
        data.get("description"),
        data.get("owner_id"),
        cia_confidentiality,
        cia_integrity,
        cia_availability,
        classification,
        security_level,
        system_id
    ))

    if cia_changed:
        cur.execute("""
            INSERT INTO system_classification_history (
                system_id,
                cia_confidentiality,
                cia_integrity,
                cia_availability,
                classification,
                security_level,
                changed_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            system_id,
            cia_confidentiality,
            cia_integrity,
            cia_availability,
            classification,
            security_level,
            session["user_id"]
        ))

    write_audit_log(
        cur,
        "UPDATE",
        "system",
        system_id,
        {
            "name": data.get("name"),
            "owner_id": data.get("owner_id"),
            "classification": classification,
            "security_level": security_level
        }
    )

    conn.commit()

    cur.close()
    conn.close()

    return {
        "message": "System updated",
        "classification": classification,
        "security_level": security_level
    }

@app.route("/continuity-page")
@require_permission("view_system")
def continuity_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template(
        "continuity.html",
        username=session["username"]
    )

@app.route("/continuity", methods=["GET"])
@require_permission("view_system")
def get_continuity_plans():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = request.args.get(
            "per_page",
            5,
            type=int
        )

        search = request.args.get(
            "search",
            "",
            type=str
        ).strip()

        if page < 1:
            page = 1

        if per_page not in (
            5,
            10,
            20,
            50
        ):
            per_page = 5

        search_condition = ""
        search_params = []

        if search:
            search_condition = """
                WHERE
                    CAST(cp.id AS TEXT) ILIKE %s
                    OR cp.name ILIKE %s
                    OR cp.description ILIKE %s
                    OR CAST(cp.rto AS TEXT) ILIKE %s
                    OR CAST(cp.rpo AS TEXT) ILIKE %s
                    OR CAST(cp.mtd AS TEXT) ILIKE %s
                    OR s.name ILIKE %s
                    OR u.username ILIKE %s
            """

            search_value = f"%{search}%"

            search_params = [
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value
            ]

        cur.execute(
            f"""
            SELECT COUNT(*)

            FROM continuity_plans cp

            LEFT JOIN systems s
                ON cp.system_id = s.system_id

            LEFT JOIN users u
                ON cp.created_by = u.user_id

            {search_condition}
            """,
            search_params
        )

        total = cur.fetchone()[0]

        pages = (
            (total + per_page - 1)
            // per_page
        )

        if pages == 0:

            page = 1

        elif page > pages:

            page = pages

        offset = (page - 1) * per_page

        cur.execute(
            f"""
            SELECT
                cp.id,
                cp.name,
                cp.description,
                cp.rto,
                cp.rpo,
                cp.mtd,
                cp.system_id,
                s.name,
                cp.created_at,
                u.username

            FROM continuity_plans cp

            LEFT JOIN systems s
                ON cp.system_id = s.system_id

            LEFT JOIN users u
                ON cp.created_by = u.user_id

            {search_condition}

            ORDER BY
                cp.created_at DESC

            LIMIT %s
            OFFSET %s
            """,
            search_params + [
                per_page,
                offset
            ]
        )

        rows = cur.fetchall()

        result = []

        for row in rows:
            result.append({
                "id":
                    row[0],
                "name":
                    row[1],
                "description":
                    row[2],
                "rto":
                    row[3],
                "rpo":
                    row[4],
                "mtd":
                    row[5],
                "system_id":
                    row[6],
                "system_name":
                    row[7],
                "created_at":
                    str(row[8])
                    if row[8]
                    else None,
                "created_by":
                    row[9]
            })

        return jsonify({
            "items":
                result,
            "page":
                page,
            "pages":
                pages,
            "per_page":
                per_page,
            "total":
                total
        })

    except Exception as e:
        print(
            "GET CONTINUITY ERROR:",
            str(e)
        )

        return jsonify({
            "error":
                "Failed to load continuity plans"
        }), 500

    finally:
        cur.close()
        conn.close()


@app.route("/continuity", methods=["POST"])
@require_permission("create_system")
def create_continuity_plan():
    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "error": "Invalid JSON data"
        }), 400

    name = str(data.get("name", "")).strip()
    description = str(data.get("description", "")).strip()

    try:
        system_id = int(data.get("system_id"))
        rto = int(data.get("rto"))
        rpo = int(data.get("rpo"))
        mtd = int(data.get("mtd"))
    except (TypeError, ValueError):
        return jsonify({
            "error": "Invalid numeric values"
        }), 400

    if not name:
        return jsonify({
            "error": "Plan name is required"
        }), 400

    if rto < 0 or rpo < 0 or mtd < 0:
        return jsonify({
            "error": "RTO, RPO and MTD cannot be negative"
        }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT 1
            FROM systems
            WHERE system_id = %s
        """, (system_id,))

        if cur.fetchone() is None:
            return jsonify({
                "error": "Information system not found"
            }), 404

        cur.execute("""
            INSERT INTO continuity_plans (
                system_id,
                name,
                description,
                rto,
                rpo,
                mtd,
                created_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            system_id,
            name,
            description,
            rto,
            rpo,
            mtd,
            session["user_id"]
        ))

        plan_id = cur.fetchone()[0]

        write_audit_log(
            cur,
            "CREATE",
            "continuity_plan",
            plan_id,
            {
                "name": data.get("name"),
                "system_id": data.get("system_id")
            }
        )

        conn.commit()

        return jsonify({
            "message": "Continuity plan created",
            "id": plan_id
        }), 201

    except Exception as e:
        conn.rollback()

        print("CREATE CONTINUITY ERROR:", str(e))

        return jsonify({
            "error": "Failed to create continuity plan"
        }), 500

    finally:
        cur.close()
        conn.close()


@app.route("/continuity/<int:plan_id>", methods=["GET"])
@require_permission("view_system")
def get_continuity_plan(plan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                id,
                name,
                description,
                rto,
                rpo,
                mtd,
                system_id
            FROM continuity_plans
            WHERE id = %s
        """, (plan_id,))

        row = cur.fetchone()

        if not row:
            return jsonify({
                "error": "Continuity plan not found"
            }), 404

        return jsonify({
            "id": row[0],
            "name": row[1],
            "description": row[2],
            "rto": row[3],
            "rpo": row[4],
            "mtd": row[5],
            "system_id": row[6]
        })

    finally:
        cur.close()
        conn.close()


@app.route("/continuity/<int:plan_id>", methods=["PUT"])
@require_permission("edit_system")
def update_continuity_plan(plan_id):
    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "error": "Invalid JSON data"
        }), 400

    name = str(data.get("name", "")).strip()
    description = str(data.get("description", "")).strip()

    try:
        system_id = int(data.get("system_id"))
        rto = int(data.get("rto"))
        rpo = int(data.get("rpo"))
        mtd = int(data.get("mtd"))
    except (TypeError, ValueError):
        return jsonify({
            "error": "Invalid numeric values"
        }), 400

    if not name:
        return jsonify({
            "error": "Plan name is required"
        }), 400

    if rto < 0 or rpo < 0 or mtd < 0:
        return jsonify({
            "error": "RTO, RPO and MTD cannot be negative"
        }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT 1
            FROM systems
            WHERE system_id = %s
        """, (system_id,))

        if cur.fetchone() is None:
            return jsonify({
                "error": "Information system not found"
            }), 404

        cur.execute("""
            UPDATE continuity_plans
            SET
                name = %s,
                description = %s,
                rto = %s,
                rpo = %s,
                mtd = %s,
                system_id = %s
            WHERE id = %s
        """, (
            name,
            description,
            rto,
            rpo,
            mtd,
            system_id,
            plan_id
        ))

        if cur.rowcount == 0:
            conn.rollback()

            return jsonify({
                "error": "Continuity plan not found"
            }), 404

        write_audit_log(
            cur,
            "UPDATE",
            "continuity_plan",
            plan_id,
            {
                "name": data.get("name"),
                "system_id": data.get("system_id")
            }
        )

        conn.commit()

        return jsonify({
            "message": "Continuity plan updated"
        })

    except Exception as e:
        conn.rollback()

        print("UPDATE CONTINUITY ERROR:", str(e))

        return jsonify({
            "error": "Failed to update continuity plan"
        }), 500

    finally:
        cur.close()
        conn.close()

@app.route("/continuity/<int:plan_id>", methods=["DELETE"])
@require_permission("delete_system")
def delete_continuity_plan(plan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            DELETE FROM continuity_plans
            WHERE id = %s
        """, (plan_id,))

        if cur.rowcount == 0:
            conn.rollback()

            return jsonify({
                "error": "Continuity plan not found"
            }), 404

        write_audit_log(
            cur,
            "DELETE",
            "continuity_plan",
            plan_id
        )

        conn.commit()

        return jsonify({
            "message": "Continuity plan deleted"
        })

    except Exception as e:
        conn.rollback()

        print("DELETE CONTINUITY ERROR:", str(e))

        return jsonify({
            "error": "Failed to delete continuity plan"
        }), 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/incidents-page")
@require_permission("view_event")
def incidents_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template("incidents.html", username=session["username"])

@app.route("/incidents", methods=["GET"])
@require_permission("view_event")
def get_incidents():
    try:
        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = request.args.get(
            "per_page",
            5,
            type=int
        )

        search = request.args.get(
            "search",
            "",
            type=str
        ).strip()

        if page < 1:
            page = 1

        if per_page not in (
            5,
            10,
            20,
            50
        ):
            per_page = 5

        conn = get_db_connection()
        cur = conn.cursor()

        search_condition = ""
        params = []

        if search:
            search_condition = """
                WHERE
                    CAST(i.incident_id AS TEXT)
                        ILIKE %s
                    OR i.name ILIKE %s
                    OR i.description ILIKE %s
                    OR i.status ILIKE %s
                    OR i.severity ILIKE %s
                    OR s.name ILIKE %s
                    OR u.username ILIKE %s
                    OR EXISTS (
                        SELECT 1
                        FROM incident_risks ir_search
                        JOIN risks r_search
                            ON r_search.risk_id = ir_search.risk_id
                        WHERE
                            ir_search.incident_id =
                                i.incident_id
                            AND r_search.name ILIKE %s
                    )
            """

            search_value = f"%{search}%"

            params = [
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value,
                search_value
            ]

        cur.execute(
            f"""
            SELECT COUNT(*)

            FROM incidents i

            LEFT JOIN systems s
                ON i.system_id = s.system_id

            LEFT JOIN users u
                ON i.created_by = u.user_id

            {search_condition}
            """,
            params
        )

        total = cur.fetchone()[0]

        pages = (
            (total + per_page - 1)
            // per_page
        )

        if pages == 0:
            page = 1

        elif page > pages:
            page = pages

        offset = (
            (page - 1)
            * per_page
        )

        cur.execute(
            f"""
            SELECT
                i.incident_id,
                i.name,
                i.description,
                i.status,
                i.severity,
                i.created_at,
                i.resolved_at,
                i.system_id,
                s.name AS system_name,
                u.username AS created_by,

                COALESCE(
                    json_agg(
                        json_build_object(
                            'id',
                            r.risk_id,
                            'name',
                            r.name
                        )
                        ORDER BY r.risk_id
                    )
                    FILTER (
                        WHERE r.risk_id IS NOT NULL
                    ),
                    '[]'::json
                ) AS risks

            FROM incidents i

            LEFT JOIN systems s
                ON i.system_id = s.system_id

            LEFT JOIN users u
                ON i.created_by = u.user_id

            LEFT JOIN incident_risks ir
                ON i.incident_id = ir.incident_id

            LEFT JOIN risks r
                ON ir.risk_id = r.risk_id

            {search_condition}

            GROUP BY
                i.incident_id,
                i.name,
                i.description,
                i.status,
                i.severity,
                i.created_at,
                i.resolved_at,
                i.system_id,
                s.name,
                u.username

            ORDER BY
                i.created_at DESC

            LIMIT %s
            OFFSET %s
            """,
            params + [
                per_page,
                offset
            ]
        )

        rows = cur.fetchall()

        result = []

        for i in rows:
            result.append({
                "id": i[0],
                "name": i[1],
                "description": i[2],
                "status": i[3],
                "severity": i[4],
                "created_at":
                    str(i[5]),
                "resolved_at":
                    str(i[6])
                    if i[6]
                    else None,
                "system_id": i[7],
                "system_name": i[8],
                "created_by": i[9],
                "risks": i[10]
            })

        cur.close()
        conn.close()

        return jsonify({
            "items": result,
            "page": page,
            "pages": pages,
            "per_page": per_page,
            "total": total
        })

    except Exception as e:
        print(
            "GET INCIDENTS ERROR:",
            str(e)
        )

        return jsonify({
            "error": "Failed to load incidents"
        }), 500

@app.route("/incidents", methods=["POST"])
@require_permission("create_event")
def create_incident():
    data = request.json or {}

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        name = data.get("name")
        description = data.get("description")
        status = data.get("status")
        system_id = data.get("system_id")
        severity = data.get("severity")
        risk_ids = data.get("risk_ids",[])

        if not isinstance(risk_ids, list):
            return {
                "error": "risk_ids must be an array"
            }, 400

        try:
            system_id = int(system_id)
        except (TypeError, ValueError):

            return {
                "error": "Invalid system_id"
            }, 400

        cleaned_risk_ids = []

        for risk_id in risk_ids:
            try:
                risk_id = int(risk_id)
            except (TypeError, ValueError):
                return {
                    "error": "Invalid risk_id"
                }, 400

            if risk_id not in cleaned_risk_ids:

                cleaned_risk_ids.append(
                    risk_id
                )

        cur.execute("""
            INSERT INTO incidents (
                name,
                description,
                status,
                system_id,
                severity,
                created_by
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            RETURNING incident_id
        """, (
            name,
            description,
            status,
            system_id,
            severity,
            session.get("user_id")
        ))

        incident_id = cur.fetchone()[0]

        for risk_id in cleaned_risk_ids:
            cur.execute("""
                INSERT INTO incident_risks (
                    incident_id,
                    risk_id
                )
                VALUES (
                    %s,
                    %s
                )
                ON CONFLICT DO NOTHING
            """, (
                incident_id,
                risk_id
            ))

        write_audit_log(
            cur,
            "CREATE",
            "incident",
            incident_id,
            {
                "name": name,
                "system_id": system_id,
                "severity": severity,
                "status": status
            }
        )

        conn.commit()

        return jsonify({
            "message": "created",
            "incident_id": incident_id,
            "risk_ids": cleaned_risk_ids
        }), 201

    except Exception as error:
        conn.rollback()

        print(
            "CREATE INCIDENT ERROR:",
            error
        )

        return {
            "error": "Failed to create incident"
        }, 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/incidents/<int:id>", methods=["GET"])
@require_permission("view_event")
def get_incident(id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                i.incident_id,
                i.name,
                i.description,
                i.system_id,
                i.severity,
                i.status,

                COALESCE(
                    json_agg(
                        ir.risk_id
                        ORDER BY ir.risk_id
                    )
                    FILTER (
                        WHERE ir.risk_id IS NOT NULL
                    ),
                    '[]'::json
                ) AS risk_ids

            FROM incidents i

            LEFT JOIN incident_risks ir
                ON i.incident_id =
                   ir.incident_id

            WHERE i.incident_id = %s

            GROUP BY
                i.incident_id,
                i.name,
                i.description,
                i.system_id,
                i.severity,
                i.status
        """, (id,))

        incident = cur.fetchone()

        if not incident:
            return {
                "error": "not found"
            }, 404

        return jsonify({
            "id": incident[0],
            "name": incident[1],
            "description": incident[2],
            "system_id": incident[3],
            "severity": incident[4],
            "status": incident[5],
            "risk_ids": incident[6]
        })


    except Exception as error:
        print(
            "GET INCIDENT ERROR:",
            error
        )

        return {
            "error": "Failed to load incident"
        }, 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/incidents/<int:id>", methods=["PUT"])
@require_permission("edit_event")
def update_incident(id):
    data = request.json or {}

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        name = data.get("name")
        description = data.get("description")
        status = data.get("status")
        system_id = data.get("system_id")
        severity = data.get("severity")
        risk_ids = data.get("risk_ids",[])

        if not isinstance(risk_ids, list):
            return {
                "error": "risk_ids must be an array"
            }, 400

        try:
            system_id = int(system_id)
        except (TypeError, ValueError):
            return {
                "error": "Invalid system_id"
            }, 400

        cleaned_risk_ids = []

        for risk_id in risk_ids:
            try:
                risk_id = int(risk_id)
            except (TypeError, ValueError):
                return {
                    "error": "Invalid risk_id"
                }, 400

            if risk_id not in cleaned_risk_ids:
                cleaned_risk_ids.append(risk_id)

        cur.execute("""
            UPDATE incidents

            SET
                name = %s,
                description = %s,
                status = %s,
                system_id = %s,
                severity = %s,

                resolved_at =
                    CASE
                        WHEN %s = 'resolved'
                        THEN NOW()
                        ELSE NULL
                    END

            WHERE incident_id = %s
        """, (
            name,
            description,
            status,
            system_id,
            severity,
            status,
            id
        ))

        if cur.rowcount == 0:
            return {
                "error": "not found"
            }, 404

        cur.execute("""
            DELETE FROM incident_risks

            WHERE incident_id = %s
        """, (id,))

        for risk_id in cleaned_risk_ids:

            cur.execute("""
                INSERT INTO incident_risks (
                    incident_id,
                    risk_id
                )
                VALUES (
                    %s,
                    %s
                )
                ON CONFLICT DO NOTHING
            """, (
                id,
                risk_id
            ))

        write_audit_log(
            cur,
            "UPDATE",
            "incident",
            id,
            {
                "name": name,
                "system_id": system_id,
                "severity": severity,
                "status": status
            }
        )

        conn.commit()

        return jsonify({
            "success": True,
            "incident_id": id,
            "risk_ids":
                cleaned_risk_ids
        })

    except Exception as error:
        conn.rollback()

        print(
            "UPDATE INCIDENT ERROR:",
            error
        )

        return {
            "error": "Failed to update incident"
        }, 500

    finally:
        cur.close()
        conn.close()
    
@app.route('/incidents/<int:id>', methods=['DELETE'])
@require_permission('delete_event')
def delete_incident(id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM incidents WHERE incident_id = %s",
        (id,)
    )

    if cur.rowcount == 0:
        conn.rollback()
        return jsonify({
            "error": "Incident not found"
        }), 404

    write_audit_log(
        cur,
        "DELETE",
        "incident",
        id
    )

    conn.commit()

    return jsonify({"success": True})

@app.route("/security-assessment-page")
@require_permission("view_reports")
def security_assessment_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template(
        "security_assessment.html",
        username=session["username"]
    )

@app.route("/security-assessments", methods=["GET"])
@require_permission("view_reports")
def get_security_assessments():
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        try:
            page = int(request.args.get("page", 1))
        except (TypeError, ValueError):
            page = 1

        try:
            per_page = int(
                request.args.get("per_page", 10)
            )
        except (TypeError, ValueError):
            per_page = 10

        search = (
            request.args.get("search", "")
            .strip()
        )

        security_level = (
            request.args.get("security_level", "")
            .strip()
        )

        if page < 1:
            page = 1

        allowed_per_page = [5, 10, 20, 50]

        if per_page not in allowed_per_page:
            per_page = 10

        conditions = []
        parameters = []

        if search:
            conditions.append("""
                (
                    CAST(sa.assessment_id AS TEXT) ILIKE %s
                    OR COALESCE(u.username, '') ILIKE %s
                    OR COALESCE(sa.security_level, '') ILIKE %s
                )
            """)

            search_value = f"%{search}%"

            parameters.extend([
                search_value,
                search_value,
                search_value
            ])

        if security_level:
            conditions.append(
                "sa.security_level = %s"
            )

            parameters.append(
                security_level
            )

        where_sql = ""

        if conditions:

            where_sql = (
                "WHERE " +
                " AND ".join(conditions)
            )

        count_query = f"""
            SELECT COUNT(*)

            FROM security_assessments sa

            LEFT JOIN users u
                ON sa.created_by = u.user_id

            {where_sql}
        """

        cur.execute(
            count_query,
            parameters
        )

        total = cur.fetchone()[0]

        pages = (
            (total + per_page - 1)
            // per_page
            if total > 0
            else 1
        )

        if page > pages:
            page = pages

        offset = (
            page - 1
        ) * per_page

        data_query = f"""
            SELECT
                sa.assessment_id,
                sa.assessment_date,

                sa.risk_management,
                sa.incident_management,
                sa.access_control,
                sa.system_protection,
                sa.continuity,

                sa.overall_score,
                sa.security_level,

                sa.systems_count,
                sa.risks_count,
                sa.active_risks_count,
                sa.incidents_count,
                sa.unresolved_incidents_count,
                sa.continuity_plans_count,

                sa.created_by,
                u.username

            FROM security_assessments sa

            LEFT JOIN users u
                ON sa.created_by = u.user_id

            {where_sql}

            ORDER BY sa.assessment_date DESC

            LIMIT %s
            OFFSET %s
        """

        data_parameters = [
            *parameters,
            per_page,
            offset
        ]

        cur.execute(
            data_query,
            data_parameters
        )

        rows = cur.fetchall()

        items = []

        for row in rows:
            items.append({
                "id": row[0],
                "assessment_date": (
                    row[1].isoformat()
                    if row[1]
                    else None
                ),
                "risk_management": row[2],
                "incident_management": row[3],
                "access_control": row[4],
                "system_protection": row[5],
                "continuity": row[6],
                "overall_score": (
                    float(row[7])
                    if row[7] is not None
                    else None
                ),
                "security_level": row[8],
                "systems_count": row[9],
                "risks_count": row[10],
                "active_risks_count": row[11],
                "incidents_count": row[12],
                "unresolved_incidents_count": row[13],
                "continuity_plans_count": row[14],
                "created_by": row[15],
                "username": row[16]
            })

        return jsonify({
            "items": items,
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": pages
        })

    except Exception as error:
        print(
            "GET SECURITY ASSESSMENTS ERROR:",
            error
        )

        return {
            "error":
                "Failed to load security assessments"
        }, 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/security-assessments/all", methods=["GET"])
@require_permission("view_reports")
def get_all_security_assessments():
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                sa.assessment_id,
                sa.assessment_date,

                sa.risk_management,
                sa.incident_management,
                sa.access_control,
                sa.system_protection,
                sa.continuity,

                sa.overall_score,
                sa.security_level,

                sa.systems_count,
                sa.risks_count,
                sa.active_risks_count,
                sa.incidents_count,
                sa.unresolved_incidents_count,
                sa.continuity_plans_count,

                sa.created_by,
                u.username

            FROM security_assessments sa

            LEFT JOIN users u
                ON sa.created_by = u.user_id

            ORDER BY sa.assessment_date ASC
        """)

        rows = cur.fetchall()

        result = []

        for row in rows:
            result.append({
                "id": row[0],
                "assessment_date": (
                    row[1].isoformat()
                    if row[1]
                    else None
                ),
                "risk_management": row[2],
                "incident_management": row[3],
                "access_control": row[4],
                "system_protection": row[5],
                "continuity": row[6],
                "overall_score": (
                    float(row[7])
                    if row[7] is not None
                    else None
                ),
                "security_level": row[8],
                "systems_count": row[9],
                "risks_count": row[10],
                "active_risks_count": row[11],
                "incidents_count": row[12],
                "unresolved_incidents_count": row[13],
                "continuity_plans_count": row[14],
                "created_by": row[15],
                "username": row[16]
            })

        return jsonify(result)

    except Exception as error:
        print(
            "GET ALL SECURITY ASSESSMENTS ERROR:",
            error
        )

        return {
            "error":
                "Failed to load security assessments"
        }, 500

    finally:
        cur.close()
        conn.close()

@app.route("/security-assessments", methods=["POST"])
@require_permission("create_reports")
def create_security_assessment():
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    data = request.json or {}

    fields = [
        "risk_management",
        "incident_management",
        "access_control",
        "system_protection",
        "continuity"
    ]

    values = {}

    for field in fields:
        value = data.get(field)

        try:
            value = int(value)
        except (TypeError, ValueError):
            return {
                "error": f"Invalid value for {field}"
            }, 400

        if value < 1 or value > 5:
            return {
                "error": f"{field} must be between 1 and 5"
            }, 400

        values[field] = value

    overall_score = round(
        (
            values["risk_management"]
            + values["incident_management"]
            + values["access_control"]
            + values["system_protection"]
            + values["continuity"]
        ) / 5,
        2
    )

    if overall_score <= 2:
        security_level = "Zems"

    elif overall_score <= 3:
        security_level = "Vidējs"

    elif overall_score <= 4:
        security_level = "Labs"

    else:
        security_level = "Augsts"

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                (SELECT COUNT(*) FROM systems),
                (SELECT COUNT(*) FROM risks),
                (SELECT COUNT(*)
                 FROM risks
                 WHERE status = 'active'),
                (SELECT COUNT(*) FROM incidents),
                (SELECT COUNT(*)
                 FROM incidents
                 WHERE status IN ('open', 'in_progress')),
                (SELECT COUNT(*) FROM continuity_plans)
        """)

        statistics = cur.fetchone()

        systems_count = statistics[0]
        risks_count = statistics[1]
        active_risks_count = statistics[2]
        incidents_count = statistics[3]
        unresolved_incidents_count = statistics[4]
        continuity_plans_count = statistics[5]

        user_id = session.get("user_id")

        cur.execute("""
            INSERT INTO security_assessments (
                risk_management,
                incident_management,
                access_control,
                system_protection,
                continuity,
                overall_score,
                security_level,
                systems_count,
                risks_count,
                active_risks_count,
                incidents_count,
                unresolved_incidents_count,
                continuity_plans_count,
                created_by
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s
            )
            RETURNING assessment_id, assessment_date
        """, (
            values["risk_management"],
            values["incident_management"],
            values["access_control"],
            values["system_protection"],
            values["continuity"],
            overall_score,
            security_level,
            systems_count,
            risks_count,
            active_risks_count,
            incidents_count,
            unresolved_incidents_count,
            continuity_plans_count,
            user_id
        ))

        result = cur.fetchone()

        write_audit_log(
            cur,
            "CREATE",
            "assessment",
            result[0],
            {
                "overall_score": float(overall_score),
                "security_level": security_level
            }
        )

        conn.commit()

        return jsonify({
            "message": "Security assessment created",
            "assessment_id": result[0],
            "assessment_date": result[1].isoformat()
                if result[1]
                else None,
            "risk_management": values["risk_management"],
            "incident_management": values["incident_management"],
            "access_control": values["access_control"],
            "system_protection": values["system_protection"],
            "continuity": values["continuity"],
            "overall_score": float(overall_score),
            "security_level": security_level,
            "systems_count": systems_count,
            "risks_count": risks_count,
            "active_risks_count": active_risks_count,
            "incidents_count": incidents_count,
            "unresolved_incidents_count": unresolved_incidents_count,
            "continuity_plans_count": continuity_plans_count
        }), 201

    except Exception as error:
        conn.rollback()

        print("CREATE SECURITY ASSESSMENT ERROR:", error)

        return {
            "error": "Failed to create security assessment"
        }, 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/security-assessments/<int:assessment_id>", methods=["PUT"])
@require_permission("edit_reports")
def update_security_assessment(assessment_id):
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    data = request.json or {}

    fields = [
        "risk_management",
        "incident_management",
        "access_control",
        "system_protection",
        "continuity"
    ]

    values = {}

    for field in fields:
        value = data.get(field)

        try:
            value = int(value)
        except (TypeError, ValueError):
            return {
                "error": f"Invalid value for {field}"
            }, 400

        if value < 1 or value > 5:
            return {
                "error": f"{field} must be between 1 and 5"
            }, 400

        values[field] = value

    overall_score = round(
        (
            values["risk_management"]
            + values["incident_management"]
            + values["access_control"]
            + values["system_protection"]
            + values["continuity"]
        ) / 5,
        2
    )

    if overall_score <= 2:
        security_level = "Zems"

    elif overall_score <= 3:
        security_level = "Vidējs"

    elif overall_score <= 4:
        security_level = "Labs"

    else:
        security_level = "Augsts"

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT assessment_id
            FROM security_assessments
            WHERE assessment_id = %s
        """, (
            assessment_id,
        ))

        if not cur.fetchone():

            return {
                "error": "Security assessment not found"
            }, 404

        cur.execute("""
            SELECT
                (SELECT COUNT(*) FROM systems),
                (SELECT COUNT(*) FROM risks),
                (SELECT COUNT(*)
                 FROM risks
                 WHERE status = 'active'),
                (SELECT COUNT(*) FROM incidents),
                (SELECT COUNT(*)
                 FROM incidents
                 WHERE status IN ('open', 'in_progress')),
                (SELECT COUNT(*) FROM continuity_plans)
        """)

        statistics = cur.fetchone()

        systems_count = statistics[0]
        risks_count = statistics[1]
        active_risks_count = statistics[2]
        incidents_count = statistics[3]
        unresolved_incidents_count = statistics[4]
        continuity_plans_count = statistics[5]

        cur.execute("""
            UPDATE security_assessments
            SET
                risk_management = %s,
                incident_management = %s,
                access_control = %s,
                system_protection = %s,
                continuity = %s,

                overall_score = %s,
                security_level = %s,

                systems_count = %s,
                risks_count = %s,
                active_risks_count = %s,
                incidents_count = %s,
                unresolved_incidents_count = %s,
                continuity_plans_count = %s

            WHERE assessment_id = %s

            RETURNING
                assessment_id,
                assessment_date
        """, (
            values["risk_management"],
            values["incident_management"],
            values["access_control"],
            values["system_protection"],
            values["continuity"],
            overall_score,
            security_level,
            systems_count,
            risks_count,
            active_risks_count,
            incidents_count,
            unresolved_incidents_count,
            continuity_plans_count,
            assessment_id
        ))

        result = cur.fetchone()

        write_audit_log(
            cur,
            "UPDATE",
            "assessment",
            assessment_id,
            {
                "overall_score": float(overall_score),
                "security_level": security_level
            }
        )

        conn.commit()

        return jsonify({
            "message": "Security assessment updated",
            "assessment_id": result[0],
            "assessment_date":
                result[1].isoformat()
                if result[1]
                else None,
            "risk_management":
                values["risk_management"],
            "incident_management":
                values["incident_management"],
            "access_control":
                values["access_control"],
            "system_protection":
                values["system_protection"],
            "continuity":
                values["continuity"],
            "overall_score":
                float(overall_score),
            "security_level":
                security_level,
            "systems_count":
                systems_count,
            "risks_count":
                risks_count,
            "active_risks_count":
                active_risks_count,
            "incidents_count":
                incidents_count,
            "unresolved_incidents_count":
                unresolved_incidents_count,
            "continuity_plans_count":
                continuity_plans_count
        }), 200

    except Exception as error:
        conn.rollback()

        print(
            "UPDATE SECURITY ASSESSMENT ERROR:",
            error
        )

        return {
            "error":
                "Failed to update security assessment"
        }, 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/security-assessments/<int:assessment_id>", methods=["DELETE"])
@require_permission("delete_reports")
def delete_security_assessment(assessment_id):
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT assessment_id
            FROM security_assessments
            WHERE assessment_id = %s
        """, (
            assessment_id,
        ))

        if not cur.fetchone():

            return {
                "error": "Security assessment not found"
            }, 404

        cur.execute("""
            DELETE FROM security_assessments
            WHERE assessment_id = %s
        """, (
            assessment_id,
        ))

        write_audit_log(
            cur,
            "DELETE",
            "assessment",
            assessment_id
        )

        conn.commit()

        return jsonify({
            "message":
                "Security assessment deleted",
            "assessment_id":
                assessment_id
        }), 200

    except Exception as error:
        conn.rollback()

        print(
            "DELETE SECURITY ASSESSMENT ERROR:",
            error
        )

        return {
            "error":
                "Failed to delete security assessment"
        }, 500

    finally:
        cur.close()
        conn.close()

@app.route("/security-assessments/<int:assessment_id>", methods=["GET"])
@require_permission("view_reports")
def get_security_assessment(assessment_id):
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                sa.assessment_id,
                sa.assessment_date,

                sa.risk_management,
                sa.incident_management,
                sa.access_control,
                sa.system_protection,
                sa.continuity,

                sa.overall_score,
                sa.security_level,

                sa.systems_count,
                sa.risks_count,
                sa.active_risks_count,
                sa.incidents_count,
                sa.unresolved_incidents_count,
                sa.continuity_plans_count,

                sa.created_by,
                u.username

            FROM security_assessments sa

            LEFT JOIN users u
                ON sa.created_by = u.user_id

            WHERE sa.assessment_id = %s
        """, (assessment_id,))

        row = cur.fetchone()

        if not row:
            return {
                "error": "Pašnovērtējums nav atrasts."
            }, 404

        result = {
            "id": row[0],
            "assessment_date": (
                row[1].isoformat()
                if row[1]
                else None
            ),
            "risk_management": row[2],
            "incident_management": row[3],
            "access_control": row[4],
            "system_protection": row[5],
            "continuity": row[6],
            "overall_score": (
                float(row[7])
                if row[7] is not None
                else None
            ),
            "security_level": row[8],
            "systems_count": row[9],
            "risks_count": row[10],
            "active_risks_count": row[11],
            "incidents_count": row[12],
            "unresolved_incidents_count": row[13],
            "continuity_plans_count": row[14],
            "created_by": row[15],
            "username": row[16]
        }

        return jsonify(result)

    except Exception as error:
        print(
            "GET SECURITY ASSESSMENT ERROR:",
            error
        )

        return {
            "error": "Failed to load security assessment"
        }, 500

    finally:
        cur.close()
        conn.close()

@app.route("/security-assessments/<int:assessment_id>/pdf", methods=["GET"])
@require_permission("view_reports")
def security_assessment_pdf(assessment_id):
    if not is_logged_in():
        return {"error": "Not authenticated"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                sa.assessment_id,
                sa.assessment_date,

                sa.risk_management,
                sa.incident_management,
                sa.access_control,
                sa.system_protection,
                sa.continuity,

                sa.overall_score,
                sa.security_level,

                sa.systems_count,
                sa.risks_count,
                sa.active_risks_count,
                sa.incidents_count,
                sa.unresolved_incidents_count,
                sa.continuity_plans_count,

                u.username

            FROM security_assessments sa

            LEFT JOIN users u
                ON sa.created_by = u.user_id

            WHERE sa.assessment_id = %s
        """, (assessment_id,))

        assessment = cur.fetchone()

        if not assessment:
            return {
                "error": "Pašnovērtējums nav atrasts."
            }, 404

        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        font_path = (
            "/usr/share/fonts/"
            "truetype/dejavu/"
            "DejaVuSans.ttf"
        )

        bold_font_path = (
            "/usr/share/fonts/"
            "truetype/dejavu/"
            "DejaVuSans-Bold.ttf"
        )

        pdfmetrics.registerFont(
            TTFont("DejaVuSans", font_path)
        )

        pdfmetrics.registerFont(
            TTFont("DejaVuSans-Bold", bold_font_path)
        )

        buffer = BytesIO()

        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        styles = getSampleStyleSheet()

        for style in styles.byName.values():
            style.fontName = "DejaVuSans"

        elements = []

        title_style = styles["Title"].clone(
            "AssessmentTitle"
        )

        title_style.fontName = "DejaVuSans-Bold"
        title_style.fontSize = 18
        title_style.leading = 22
        title_style.alignment = TA_CENTER

        heading_style = styles["Heading2"].clone(
            "AssessmentHeading"
        )

        heading_style.fontName = "DejaVuSans-Bold"
        heading_style.fontSize = 13
        heading_style.leading = 16

        normal_style = styles["Normal"].clone(
            "AssessmentNormal"
        )

        normal_style.fontName = "DejaVuSans"
        normal_style.fontSize = 9
        normal_style.leading = 12

        elements = []

        elements.append(
            Paragraph(
                "Informācijas drošības pašnovērtējums",
                title_style
            )
        )

        elements.append(
            Spacer(1, 20)
        )

        assessment_date = (
            assessment[1].strftime("%d.%m.%Y %H:%M")
            if assessment[1]
            else "-"
        )

        elements.append(
            Paragraph(
                f"<b>Datums:</b> {assessment_date}",
                normal_style
            )
        )

        elements.append(
            Paragraph(
                f"<b>Izveidoja:</b> "
                f"{assessment[15] or '-'}",
                normal_style
            )
        )

        elements.append(
            Spacer(1, 20)
        )

        elements.append(
            Paragraph(
                "Pašnovērtējuma rezultāti",
                heading_style
            )
        )

        elements.append(
            Spacer(1, 10)
        )

        assessment_table_data = [
            [
                "Joma",
                "Vērtējums"
            ],

            [
                "Risku pārvaldība",
                str(assessment[2])
            ],

            [
                "Incidentu pārvaldība",
                str(assessment[3])
            ],

            [
                "Piekļuves kontrole",
                str(assessment[4])
            ],

            [
                "Sistēmu aizsardzība",
                str(assessment[5])
            ],

            [
                "Darbības nepārtrauktība",
                str(assessment[6])
            ]
        ]

        assessment_table = Table(
            assessment_table_data,
            colWidths=[350, 100]
        )

        assessment_table.setStyle(
            TableStyle([
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    "DejaVuSans"
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "DejaVuSans-Bold"
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1e293b")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )
            ])
        )

        elements.append(
            assessment_table
        )

        elements.append(
            Spacer(1, 20)
        )

        elements.append(
            Paragraph(
                "Kopējais rezultāts",
                heading_style
            )
        )

        elements.append(
            Spacer(1, 10)
        )

        overall_score = (
            float(assessment[7])
            if assessment[7] is not None
            else 0
        )

        summary_data = [
            [
                "Kopējais vērtējums",
                f"{overall_score:.2f}"
            ],
            [
                "Drošības līmenis",
                assessment[8] or "-"
            ]
        ]

        summary_table = Table(
            summary_data,
            colWidths=[350, 100]
        )

        summary_table.setStyle(
            TableStyle([
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    "DejaVuSans"
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (0, -1),
                    "DejaVuSans-Bold"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#e2e8f0")
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )
            ])
        )

        elements.append(
            summary_table
        )

        elements.append(
            Spacer(1, 20)
        )

        elements.append(
            Paragraph(
                "Sistēmas stāvokļa kopsavilkums",
                heading_style
            )
        )

        elements.append(
            Spacer(1, 10)
        )

        statistics_data = [
            [
                "Rādītājs",
                "Skaits"
            ],
            [
                "Sistēmas",
                str(assessment[9])
            ],
            [
                "Riski",
                str(assessment[10])
            ],
            [
                "Aktīvie riski",
                str(assessment[11])
            ],
            [
                "Incidenti",
                str(assessment[12])
            ],
            [
                "Neatrisinātie incidenti",
                str(assessment[13])
            ],
            [
                "Nepārtrauktības plāni",
                str(assessment[14])
            ]
        ]

        statistics_table = Table(
            statistics_data,
            colWidths=[350, 100]
        )

        statistics_table.setStyle(
            TableStyle([
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    "DejaVuSans"
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "DejaVuSans-Bold"
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#1e293b")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )
            ])
        )

        elements.append(
            statistics_table
        )

        document.build(elements)

        buffer.seek(0)

        return (
            buffer.getvalue(),
            200,
            {
                "Content-Type": "application/pdf",
                "Content-Disposition":
                    f"attachment; "
                    f"filename=security_assessment_"
                    f"{assessment_id}.pdf"
            }
        )

    except Exception as error:
        print(
            "SECURITY ASSESSMENT PDF ERROR:",
            error
        )

        return {
            "error":
                "Neizdevās izveidot PDF pārskatu."
        }, 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/security-assessments/<int:assessment_id>/analysis", methods=["GET"])
@require_permission("view_reports")
def get_security_assessment_analysis(assessment_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                assessment_id,
                risk_management,
                incident_management,
                access_control,
                system_protection,
                continuity
            FROM security_assessments
            WHERE assessment_id = %s
        """, (assessment_id,))

        assessment = cur.fetchone()

        if not assessment:
            return {
                "error": "Pašnovērtējums nav atrasts."
            }, 404

        areas = [
            {
                "key": "risk_management",
                "name": "Risku pārvaldība",
                "score": assessment[1]
            },
            {
                "key": "incident_management",
                "name": "Incidentu pārvaldība",
                "score": assessment[2]
            },
            {
                "key": "access_control",
                "name": "Piekļuves kontrole",
                "score": assessment[3]
            },
            {
                "key": "system_protection",
                "name": "Sistēmu aizsardzība",
                "score": assessment[4]
            },
            {
                "key": "continuity",
                "name": "Darbības nepārtrauktība",
                "score": assessment[5]
            }
        ]

        for area in areas:
            score = area["score"]

            if score <= 2:
                area["status"] = "attention"
            elif score == 3:
                area["status"] = "medium"
            else:
                area["status"] = "good"

        areas.sort(
            key=lambda x: x["score"]
        )

        attention_areas = [
            area
            for area in areas
            if area["score"] <= 2
        ]

        return jsonify({
            "assessment_id": assessment[0],
            "areas": areas,
            "attention_areas": attention_areas,
            "attention_count": len(attention_areas)
        })

    except Exception as error:
        print(
            "SECURITY ASSESSMENT ANALYSIS ERROR:",
            error
        )

        return {
            "error":
                "Neizdevās analizēt pašnovērtējuma rezultātus."
        }, 500

    finally:
        cur.close()
        conn.close()

@app.route("/reports-page")
@require_permission("view_reports")
def reports_page():
    if not is_logged_in():
        return redirect("/login")

    return render_template(
        "reports.html",
        username=session["username"]
    )
    
@app.route("/reports/summary", methods=["GET"])
@require_permission("view_reports")
def reports_summary():
    if not is_logged_in():
        return {
            "error": "Not authenticated"
        }, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT COUNT(*)
            FROM systems
        """)

        systems_count = cur.fetchone()[0]

        cur.execute("""
            SELECT
                COUNT(*),

                COUNT(*) FILTER (
                    WHERE status = 'active'
                ),

                COUNT(*) FILTER (
                    WHERE risk_level <= 5
                ),

                COUNT(*) FILTER (
                    WHERE risk_level > 5
                    AND risk_level <= 12
                ),

                COUNT(*) FILTER (
                    WHERE risk_level > 12
                )

            FROM risks
        """)

        risk_row = cur.fetchone()

        risks_count = risk_row[0]
        active_risks_count = risk_row[1]

        low_risks_count = risk_row[2]
        medium_risks_count = risk_row[3]
        high_risks_count = risk_row[4]

        cur.execute("""
            SELECT
                COUNT(*),
                COUNT(*) FILTER (
                    WHERE status != 'resolved'
                ),
                COUNT(*) FILTER (
                    WHERE severity = 'low'
                ),
                COUNT(*) FILTER (
                    WHERE severity = 'medium'
                ),
                COUNT(*) FILTER (
                    WHERE severity = 'high'
                )
            FROM incidents
        """)

        incident_row = cur.fetchone()

        incidents_count = incident_row[0]
        unresolved_incidents_count = incident_row[1]

        low_incidents_count = incident_row[2]
        medium_incidents_count = incident_row[3]
        high_incidents_count = incident_row[4]

        cur.execute("""
            SELECT COUNT(*)
            FROM continuity_plans
        """)

        continuity_plans_count = cur.fetchone()[0]

        cur.execute("""
            SELECT
                assessment_id,
                assessment_date,
                overall_score,
                security_level
            FROM security_assessments
            ORDER BY assessment_date DESC
            LIMIT 1
        """)

        assessment = cur.fetchone()

        latest_assessment = None

        if assessment:
            latest_assessment = {
                "id":
                    assessment[0],
                "date":
                    str(assessment[1]),
                "overall_score":
                    float(assessment[2]),
                "security_level":
                    assessment[3]
            }

        return jsonify({
            "systems": {
                "total":
                    systems_count
            },
            "risks": {

                "total":
                    risks_count,

                "active":
                    active_risks_count,

                "low":
                    low_risks_count,

                "medium":
                    medium_risks_count,

                "high":
                    high_risks_count
            },
            "incidents": {

                "total":
                    incidents_count,

                "unresolved":
                    unresolved_incidents_count,

                "low":
                    low_incidents_count,

                "medium":
                    medium_incidents_count,

                "high":
                    high_incidents_count,

            },
            "continuity_plans": {

                "total":
                    continuity_plans_count
            },
            "latest_assessment":
                latest_assessment
        })

    except Exception as error:
        print(
            "REPORTS SUMMARY ERROR:",
            str(error)
        )

        return jsonify({
            "error": "Failed to load report summary"
        }), 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/reports/risk-dynamics", methods=["GET"])
@require_permission("view_reports")
def reports_risk_dynamics():
    if not is_logged_in():
        return {
            "error": "Not authenticated"
        }, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            WITH days AS (
                SELECT generate_series(
                    CURRENT_DATE - INTERVAL '29 days',
                    CURRENT_DATE,
                    INTERVAL '1 day'
                )::date AS day
            ),

            deleted_risks AS (
                SELECT
                    c.entity_id,
                    c.created_at AS created_at,
                    MIN(d.created_at) AS deleted_at
                FROM audit_log c
                JOIN audit_log d
                    ON d.entity_type = 'risk'
                    AND d.entity_id = c.entity_id
                    AND d.action = 'DELETE'
                    AND d.created_at > c.created_at
                WHERE
                    c.entity_type = 'risk'
                    AND c.action = 'CREATE'
                GROUP BY
                    c.entity_id,
                    c.created_at
            )

            SELECT
                TO_CHAR(days.day, 'YYYY-MM-DD') AS day,

                (
                    SELECT COUNT(*)
                    FROM risks r
                    WHERE r.created_at < days.day + INTERVAL '1 day'
                )

                +

                (
                    SELECT COUNT(*)
                    FROM deleted_risks dr
                    WHERE
                        dr.created_at < days.day + INTERVAL '1 day'
                        AND dr.deleted_at >= days.day + INTERVAL '1 day'
                )

                AS count

            FROM days

            ORDER BY days.day
        """)

        rows = cur.fetchall()

        result = []

        for row in rows:
            result.append({
                "date": row[0],
                "count": row[1]
            })

        return jsonify({
            "items": result
        })

    except Exception as error:
        print(
            "REPORTS RISK DYNAMICS ERROR:",
            str(error)
        )

        return jsonify({
            "error":
                "Neizdevās ielādēt risku dinamiku."
        }), 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/reports/assessment-dynamics", methods=["GET"])
@require_permission("view_reports")
def reports_assessment_dynamics():
    if not is_logged_in():
        return {
            "error": "Not authenticated"
        }, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                assessment_date,
                overall_score,
                security_level
            FROM security_assessments
            ORDER BY assessment_date ASC
        """)

        rows = cur.fetchall()

        result = []

        for row in rows:
            result.append({
                "date": row[0].strftime("%Y-%m-%d"),
                "score": float(row[1]),
                "security_level": row[2]
            })

        return jsonify({
            "items": result
        })

    except Exception as error:
        print(
            "REPORTS ASSESSMENT DYNAMICS ERROR:",
            str(error)
        )

        return jsonify({
            "error":
                "Neizdevās ielādēt pašnovērtējuma dinamiku."
        }), 500

    finally:
        cur.close()
        conn.close()
        
@app.route("/reports/assessment-focus", methods=["GET"])
@require_permission("view_reports")
def reports_assessment_focus():
    if not is_logged_in():
        return {
            "error": "Not authenticated"
        }, 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                risk_management,
                incident_management,
                access_control,
                system_protection,
                continuity
            FROM security_assessments
            ORDER BY assessment_date DESC
            LIMIT 1
        """)

        row = cur.fetchone()

        if not row:
            return jsonify({
                "items": []
            })

        areas = [
            {
                "name": "Risku pārvaldība",
                "score": float(row[0])
            },
            {
                "name": "Incidentu pārvaldība",
                "score": float(row[1])
            },
            {
                "name": "Piekļuves kontrole",
                "score": float(row[2])
            },
            {
                "name": "Sistēmu aizsardzība",
                "score": float(row[3])
            },
            {
                "name": "Darbības nepārtrauktība",
                "score": float(row[4])
            }
        ]

        areas.sort(
            key=lambda item: item["score"]
        )

        for area in areas:
            if area["score"] <= 2:
                area["status"] = "Nepieciešama uzmanība"

            elif area["score"] == 3:
                area["status"] = "Vidējs līmenis"

            else:
                area["status"] = "Labs līmenis"

        return jsonify({
            "items": areas
        })

    except Exception as error:
        print(
            "REPORTS ASSESSMENT FOCUS ERROR:",
            str(error)
        )

        return jsonify({
            "error":
                "Neizdevās ielādēt jomas, kurām nepieciešama uzmanība."
        }), 500

    finally:
        cur.close()
        conn.close()
        
def build_report_data(report_type):
    allowed_types = {
        "full",
        "risks",
        "incidents",
        "assessment"
    }

    if report_type not in allowed_types:
        raise ValueError(
            "Nezināms atskaites tips."
        )

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        report = {
            "type": report_type,
            "generated_at": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        }

        if report_type == "full":
            cur.execute("""
                SELECT COUNT(*)
                FROM systems
            """)

            systems_total = cur.fetchone()[0]

            report["systems"] = {
                "total": systems_total
            }

        if report_type in ("full", "risks"):
            cur.execute("""
                SELECT
                    COUNT(*) AS total,

                    COUNT(*) FILTER (
                        WHERE status = 'active'
                    ) AS active,

                    COUNT(*) FILTER (
                        WHERE risk_level <= 5
                    ) AS low,

                    COUNT(*) FILTER (
                        WHERE risk_level BETWEEN 6 AND 12
                    ) AS medium,

                    COUNT(*) FILTER (
                        WHERE risk_level >= 13
                    ) AS high

                FROM risks
            """)

            row = cur.fetchone()

            report["risks"] = {
                "total": row[0],
                "active": row[1],
                "low": row[2],
                "medium": row[3],
                "high": row[4]
            }

        if report_type in ("full", "incidents"):
            cur.execute("""
                SELECT
                    COUNT(*) AS total,

                    COUNT(*) FILTER (
                        WHERE status != 'resolved'
                    ) AS unresolved,

                    COUNT(*) FILTER (
                        WHERE severity = 'low'
                    ) AS low,

                    COUNT(*) FILTER (
                        WHERE severity = 'medium'
                    ) AS medium,

                    COUNT(*) FILTER (
                        WHERE severity = 'high'
                    ) AS high

                FROM incidents
            """)

            row = cur.fetchone()

            report["incidents"] = {
                "total": row[0],
                "unresolved": row[1],
                "low": row[2],
                "medium": row[3],
                "high": row[4]
            }

        if report_type == "full":
            cur.execute("""
                SELECT COUNT(*)
                FROM continuity_plans
            """)

            continuity_total = cur.fetchone()[0]

            report["continuity_plans"] = {
                "total": continuity_total
            }

        if report_type in ("full", "assessment"):
            cur.execute("""
                SELECT
                    assessment_date,
                    overall_score,
                    security_level

                FROM security_assessments

                ORDER BY assessment_date DESC

                LIMIT 1
            """)

            row = cur.fetchone()

            if row:
                report["latest_assessment"] = {
                    "date": row[0].strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                    "overall_score": float(row[1]),
                    "security_level": row[2]
                }

            else:
                report["latest_assessment"] = None

        return report

    finally:
        cur.close()
        conn.close()
        
@app.route("/reports/generate", methods=["POST"])
@require_permission("view_reports")
def generate_report():
    if not is_logged_in():
        return jsonify({
            "error": "Not authenticated"
        }), 401

    data = request.get_json(silent=True) or {}

    report_type = data.get(
        "report_type",
        "full"
    )

    try:
        report = build_report_data(
            report_type
        )

        return jsonify(report)

    except ValueError as error:
        return jsonify({
            "error": str(error)
        }), 400

    except Exception as error:
        import traceback

        print(
            "REPORT GENERATION ERROR:",
            str(error)
        )

        traceback.print_exc()

        return jsonify({
            "error": "Neizdevās izveidot atskaiti."
        }), 500

@app.route("/reports/export/pdf", methods=["POST"])
@require_permission("view_reports")
def export_report_pdf():
    if not is_logged_in():
        return jsonify({
            "error": "Not authenticated"
        }), 401

    data = request.get_json(
        silent=True
    ) or {}

    report_type = data.get(
        "report_type",
        "full"
    )

    try:
        report = build_report_data(
            report_type
        )

        buffer = io.BytesIO()

        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        styles = getSampleStyleSheet()

        for style in styles.byName.values():
            style.fontName = "DejaVuSans"

        elements = []

        elements = []

        elements.append(
            Paragraph(
                "Kiberdrošības analītiskais pārskats",
                styles["Title"]
            )
        )

        elements.append(
            Spacer(1, 10)
        )

        elements.append(
            Paragraph(
                f"Izveidots: "
                f"{report['generated_at']}",
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 20)
        )

        if "systems" in report:
            elements.append(
                Paragraph(
                    "Sistēmas",
                    styles["Heading2"]
                )
            )

            table = Table([
                ["Rādītājs", "Skaits"],
                [
                    "Kopējais sistēmu skaits",
                    report["systems"]["total"]
                ]
            ])

            table.setStyle(
                TableStyle([
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "DejaVuSans"
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e293b")
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ])
            )

            elements.append(table)

            elements.append(
                Spacer(1, 20)
            )

        if "risks" in report:
            risks = report["risks"]

            elements.append(
                Paragraph(
                    "Riski",
                    styles["Heading2"]
                )
            )

            table = Table([
                ["Rādītājs", "Skaits"],
                ["Kopā", risks["total"]],
                ["Aktīvi", risks["active"]],
                ["Zems", risks["low"]],
                ["Vidējs", risks["medium"]],
                ["Augsts", risks["high"]]
            ])

            table.setStyle(
                TableStyle([
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "DejaVuSans"
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e293b")
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ])
            )

            elements.append(table)

            elements.append(
                Spacer(1, 20)
            )

        if "incidents" in report:
            incidents = report["incidents"]

            elements.append(
                Paragraph(
                    "Incidenti",
                    styles["Heading2"]
                )
            )

            table = Table([
                ["Rādītājs", "Skaits"],
                ["Kopā", incidents["total"]],
                [
                    "Neatrisināti",
                    incidents["unresolved"]
                ],
                ["Zema smaguma", incidents["low"]],
                ["Vidēja smaguma", incidents["medium"]],
                ["Augsta smaguma", incidents["high"]]
            ])

            table.setStyle(
                TableStyle([
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "DejaVuSans"
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e293b")
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ])
            )

            elements.append(table)

            elements.append(
                Spacer(1, 20)
            )

        if "continuity_plans" in report:
            elements.append(
                Paragraph(
                    "Darbības nepārtrauktība",
                    styles["Heading2"]
                )
            )

            table = Table([
                ["Rādītājs", "Skaits"],
                [
                    "Nepārtrauktības plāni",
                    report[
                        "continuity_plans"
                    ]["total"]
                ]
            ])

            table.setStyle(
                TableStyle([
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "DejaVuSans"
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e293b")
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ])
            )

            elements.append(table)

            elements.append(
                Spacer(1, 20)
            )

        if "latest_assessment" in report:
            assessment = report[
                    "latest_assessment"
                ]

            elements.append(
                Paragraph(
                    "Pēdējais pašnovērtējums",
                    styles["Heading2"]
                )
            )

            if assessment:
                table = Table([
                    [
                        "Rādītājs",
                        "Vērtība"
                    ],
                    [
                        "Kopējais vērtējums",
                        assessment[
                            "overall_score"
                        ]
                    ],
                    [
                        "Drošības līmenis",
                        assessment[
                            "security_level"
                        ]
                    ],
                    [
                        "Datums",
                        assessment["date"]
                    ]
                ])

            else:
                table = Table([
                    [
                        "Pašnovērtējums"
                    ],
                    [
                        "Nav pieejamu datu"
                    ]
                ])

            table.setStyle(
                TableStyle([
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "DejaVuSans"
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1e293b")
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    )
                ])
            )

            elements.append(table)

        document.build(elements)

        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=
                "kiberdrosibas_atskaite.pdf"
        )

    except ValueError as error:
        return jsonify({
            "error": str(error)
        }), 400

    except Exception as error:
        import traceback

        print(
            "PDF EXPORT ERROR:",
            str(error)
        )

        traceback.print_exc()

        return jsonify({
            "error":
                "Neizdevās eksportēt PDF."
        }), 500
        
@app.route("/reports/export/excel",methods=["POST"])
@require_permission("view_reports")
def export_report_excel():
    if not is_logged_in():
        return jsonify({
            "error": "Not authenticated"
        }), 401

    data = request.get_json(
        silent=True
    ) or {}

    report_type = data.get(
        "report_type",
        "full"
    )

    try:
        report = build_report_data(
            report_type
        )

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Atskaite"

        sheet["A1"] = (
            "Kiberdrošības analītiskais pārskats"
        )

        sheet["A1"].font = Font(
            bold=True,
            size=16
        )

        sheet["A2"] = "Izveidots"

        sheet["B2"] = (
            report["generated_at"]
        )

        row = 4

        if "systems" in report:
            sheet.cell(
                row=row,
                column=1,
                value="Sistēmas"
            ).font = Font(
                bold=True
            )

            row += 1

            sheet.cell(
                row=row,
                column=1,
                value="Kopējais sistēmu skaits"
            )

            sheet.cell(
                row=row,
                column=2,
                value=report[
                    "systems"
                ]["total"]
            )

            row += 2

        if "risks" in report:
            risks = report["risks"]

            sheet.cell(
                row=row,
                column=1,
                value="Riski"
            ).font = Font(
                bold=True
            )

            row += 1

            risk_rows = [
                (
                    "Kopā",
                    risks["total"]
                ),
                (
                    "Aktīvi",
                    risks["active"]
                ),
                (
                    "Zems",
                    risks["low"]
                ),
                (
                    "Vidējs",
                    risks["medium"]
                ),
                (
                    "Augsts",
                    risks["high"]
                )
            ]

            for name, value in risk_rows:
                sheet.cell(
                    row=row,
                    column=1,
                    value=name
                )

                sheet.cell(
                    row=row,
                    column=2,
                    value=value
                )

                row += 1

            row += 1

        if "incidents" in report:
            incidents = report["incidents"]

            sheet.cell(
                row=row,
                column=1,
                value="Incidenti"
            ).font = Font(
                bold=True
            )

            row += 1

            incident_rows = [
                (
                    "Kopā",
                    incidents["total"]
                ),
                (
                    "Neatrisināti",
                    incidents["unresolved"]
                ),
                (
                    "Zema smaguma",
                    incidents["low"]
                ),
                (
                    "Vidēja smaguma",
                    incidents["medium"]
                ),
                (
                    "Augsta smaguma",
                    incidents["high"]
                )
            ]

            for name, value in incident_rows:
                sheet.cell(
                    row=row,
                    column=1,
                    value=name
                )

                sheet.cell(
                    row=row,
                    column=2,
                    value=value
                )

                row += 1

            row += 1

        if "continuity_plans" in report:
            sheet.cell(
                row=row,
                column=1,
                value="Darbības nepārtrauktība"
            ).font = Font(
                bold=True
            )

            row += 1

            sheet.cell(
                row=row,
                column=1,
                value="Nepārtrauktības plāni"
            )

            sheet.cell(
                row=row,
                column=2,
                value=report[
                    "continuity_plans"
                ]["total"]
            )

            row += 2

        if "latest_assessment" in report:
            assessment = report[
                    "latest_assessment"
                ]

            sheet.cell(
                row=row,
                column=1,
                value="Pēdējais pašnovērtējums"
            ).font = Font(
                bold=True
            )

            row += 1

            if assessment:
                assessment_rows = [
                    (
                        "Kopējais vērtējums",
                        assessment[
                            "overall_score"
                        ]
                    ),
                    (
                        "Drošības līmenis",
                        assessment[
                            "security_level"
                        ]
                    ),
                    (
                        "Datums",
                        assessment[
                            "date"
                        ]
                    )
                ]

                for name, value in assessment_rows:
                    sheet.cell(
                        row=row,
                        column=1,
                        value=name
                    )

                    sheet.cell(
                        row=row,
                        column=2,
                        value=value
                    )

                    row += 1

            else:
                sheet.cell(
                    row=row,
                    column=1,
                    value="Pašnovērtējums"
                )

                sheet.cell(
                    row=row,
                    column=2,
                    value="Nav pieejamu datu"
                )

        sheet.column_dimensions[
            "A"
        ].width = 32

        sheet.column_dimensions[
            "B"
        ].width = 28

        buffer = io.BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        return send_file(
            buffer,
            mimetype=
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=
                "kiberdrosibas_atskaite.xlsx"
        )

    except ValueError as error:
        return jsonify({
            "error": str(error)
        }), 400

    except Exception as error:
        import traceback

        print(
            "EXCEL EXPORT ERROR:",
            str(error)
        )

        traceback.print_exc()

        return jsonify({
            "error":
                "Neizdevās eksportēt Excel."
        }), 500

@app.route("/admin-page")
@require_permission("manage_users")
def admin_page():
    return render_template("admin.html")

@app.route("/admin/users")
@require_permission("manage_users")
def admin_users():
    return render_template("admin_users.html")
    
@app.route("/admin/system-status")
@require_permission("manage_users")
def admin_system_status():
    return render_template("admin_system.html")

@app.route("/api/users", methods=["GET", "POST"])
@require_permission("manage_users")
def users():
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "GET":
        cur.execute("""
            SELECT u.user_id, u.username, r.name, u.created_at
            FROM users u
            JOIN roles r ON u.role_id = r.id
        """)
        rows = cur.fetchall()

        users = []
        for u in rows:
            users.append({
                "user_id": u[0],
                "username": u[1],
                "role": u[2],
                "created_at": str(u[3])
            })

        cur.close()
        conn.close()
        return jsonify(users)

    if request.method == "POST":
        data = request.json

        username = data.get("username")
        password = data.get("password")
        role = data.get("role", "guest")

        if not username or not password:
            return {"error": "Missing fields"}, 400

        password_hash = generate_password_hash(password)

        try:
            cur.execute("SELECT 1 FROM users WHERE username = %s", (username,))
            if cur.fetchone():
                return {"error": "User already exists"}, 400

            cur.execute("""
                INSERT INTO users (username, password_hash, role_id)
                VALUES (%s, %s,
                    (SELECT id FROM roles WHERE name = %s)
                )
                RETURNING user_id
            """, (username, password_hash, role))

            user_id = cur.fetchone()[0]

            write_audit_log(
                cur,
                "CREATE",
                "user",
                user_id,
                {
                    "username": username,
                    "role": role
                }
            )

            conn.commit()

            return {"message": "User created"}

        except Exception as e:
            print("CREATE USER ERROR:", str(e))
            return {"error": "Failed to create user"}, 500

        finally:
            cur.close()
            conn.close()

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@require_permission("manage_users")
def delete_user(user_id):
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT username FROM users WHERE user_id = %s", (user_id,))
    user = cur.fetchone()

    if not user:
        return {"error": "User not found"}, 404

    if user[0] == session.get("username"):
        return {"error": "You cannot delete yourself"}, 400

    cur.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
    
    write_audit_log(
        cur,
        "DELETE",
        "user",
        user_id,
        {
            "username": user[0]
        }
    )
    
    conn.commit()

    cur.close()
    conn.close()

    return {"message": "User deleted"}
    
@app.route("/api/users/<int:user_id>/role", methods=["PUT"])
@require_permission("manage_users")
def update_user_role(user_id):
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    data = request.json
    new_role = data.get("role")

    if not new_role:
        return {"error": "Role required"}, 400

    conn = get_db_connection()
    cur = conn.cursor()

    if user_id == session.get("user_id"):
        return {"error": "Cannot change your own role"}, 400
    
    cur.execute("""
        SELECT u.username, r.name
        FROM users u
        JOIN roles r ON u.role_id = r.id
        WHERE u.user_id = %s
    """, (user_id,))

    user = cur.fetchone()

    if not user:
        cur.close()
        conn.close()
        return {"error": "User not found"}, 404

    username = user[0]
    old_role = user[1]
    
    cur.execute("SELECT id FROM roles WHERE name = %s", (new_role,))
    role = cur.fetchone()

    if not role:
        return {"error": "Invalid role"}, 400

    cur.execute("""
        UPDATE users
        SET role_id = %s
        WHERE user_id = %s
    """, (role[0], user_id))

    write_audit_log(
        cur,
        "UPDATE",
        "user",
        user_id,
        {
            "username": username,
            "old_role": old_role,
            "new_role": new_role
        }
    )

    conn.commit()

    cur.close()
    conn.close()

    return {"message": "Role updated"}
    
@app.route("/api/roles", methods=["GET"])
@require_permission("manage_users")
def roles():
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT id, name FROM roles")
    roles = cur.fetchall()

    cur.execute("SELECT name FROM permissions")
    permissions = [p[0] for p in cur.fetchall()]

    cur.close()
    conn.close()

    return jsonify({
        "roles": [{"id": r[0], "name": r[1]} for r in roles],
        "permissions": permissions
    })
    
@app.route("/api/roles/<int:role_id>", methods=["GET"])
@require_permission("manage_users")
def get_role(role_id):
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT p.name
        FROM role_permissions rp
        JOIN permissions p ON rp.permission_id = p.id
        WHERE rp.role_id = %s
    """, (role_id,))

    permissions = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return jsonify({"permissions": permissions})
    
@app.route("/api/roles/<int:role_id>", methods=["PUT"])
@require_permission("manage_users")
def update_role(role_id):
    if not has_permission("manage_users"):
        return {"error": "Forbidden"}, 403

    data = request.json
    permissions = data.get("permissions", [])

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT name
            FROM roles
            WHERE id = %s
        """, (role_id,))

        role = cur.fetchone()

        if not role:
            return {"error": "Role not found"}, 404

        role_name = role[0]

        cur.execute("""
            SELECT p.name
            FROM role_permissions rp
            JOIN permissions p
                ON rp.permission_id = p.id
            WHERE rp.role_id = %s
        """, (role_id,))

        old_permissions = {
            row[0]
            for row in cur.fetchall()
        }

        new_permissions = set(permissions)

        permissions_added = sorted(
            new_permissions - old_permissions
        )

        permissions_removed = sorted(
            old_permissions - new_permissions
        )

        cur.execute("""
            DELETE FROM role_permissions
            WHERE role_id = %s
        """, (role_id,))

        for perm in new_permissions:
            cur.execute("""
                INSERT INTO role_permissions (
                    role_id,
                    permission_id
                )
                SELECT %s, id
                FROM permissions
                WHERE name = %s
            """, (role_id, perm))

        if permissions_added or permissions_removed:
            write_audit_log(
                cur,
                "UPDATE",
                "role",
                role_id,
                {
                    "role": role_name,
                    "permissions_added": permissions_added,
                    "permissions_removed": permissions_removed
                }
            )

        conn.commit()

        return {"message": "updated"}

    except Exception as e:
        conn.rollback()

        print("UPDATE ROLE ERROR:", str(e))

        return {
            "error": "Failed to update role"
        }, 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/admin/roles")
@require_permission("manage_users")
def admin_roles():
    if not is_logged_in():
        return redirect("/login")

    if not has_permission("manage_users"):
        return "Forbidden", 403

    return render_template("admin_roles.html")
    
@app.route("/api/me", methods=["GET"])
def get_me():
    if "user_id" not in session:
        return {"error": "Unauthorized"}, 401

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT r.name
        FROM users u
        JOIN roles r ON u.role_id = r.id
        WHERE u.user_id = %s
    """, (session["user_id"],))

    role = cur.fetchone()

    cur.execute("""
        SELECT p.name
        FROM users u
        JOIN roles r ON u.role_id = r.id
        JOIN role_permissions rp ON rp.role_id = r.id
        JOIN permissions p ON rp.permission_id = p.id
        WHERE u.user_id = %s
    """, (session["user_id"],))

    permissions = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return {
        "user_id": session["user_id"],
        "username": session.get("username"),
        "role": role[0] if role else None,
        "permissions": permissions
    }

@app.route("/logout")
def logout():
    user_id = session.get("user_id")

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        if user_id:
            cur.execute("""
                UPDATE users
                SET session_version = session_version + 1
                WHERE user_id = %s
            """, (user_id,))

            write_audit_log(
                cur,
                "LOGOUT",
                "authentication",
                user_id
            )

            conn.commit()

    except Exception as e:
        conn.rollback()
        print("LOGOUT AUDIT ERROR:", str(e))

    finally:
        cur.close()
        conn.close()

    session.clear()

    return redirect("/login")